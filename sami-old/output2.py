import datetime
import numpy as np
import pandas as pd
import math
import re
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook

global e1
global e2
global button1
global button2

def deb_fun():
    global debut
    global date_deb
    debut = e1.get()
    #date_deb = e1.get()

    if (debut == ""):
        messagebox.showerror("", "Veuillez entrer une date debut")

    else:
        try:
            #date_deb = datetime.datetime.strptime(debut, "%d/%m/%y").strftime('%d/%m/%y')
            date_deb = datetime.datetime.strptime(debut, "%d/%m/%y")
        except:
            messagebox.showerror("", "Veuillez suivre ce format: jj/mm/aa")
            e1.delete(0,"end")
        #print(date_deb)
        label3 = Label(root, text="La date debut est %s" %(date_deb.strftime('%d/%m/%y')), font=("Arial", 13))
        label3.place(x=10, y=50)
        e1.delete(0,"end")
        e1.config(state='disabled')
        e2.config(state='normal')
        button1.config(state='disabled')
        button2.config(state='normal')
        #date_debut = date_deb
    
def fin_fun():
    global fin
    global date_fin
    fin = e2.get()

    if (fin == ""):
        messagebox.showerror("", "Veuillez entrer une date fin")
    else:
        try:
            date_fin = datetime.datetime.strptime(fin, "%d/%m/%y")
            if (date_fin < date_deb):
                messagebox.showerror("", "Erreur: date fin < date debut")
                return
        except:
            messagebox.showerror("", "Veuillez suivre ce format: dd/mm/yy")
            e2.delete(0,"end")
            return
        label4 = Label(root, text="La date fin est %s" %(date_fin.strftime('%d/%m/%y')), font=("Arial", 13))
        label4.place(x=10, y=130)
        e2.config(state='disabled')
        button2.config(state='disabled')
        messagebox.showinfo("Output", "Veuillez cliquer pour voir les resultats suivants sur le meme dossier:\n1. CA_BejaouiS_SaidiA\n2. Commissions_BejaouiS_SaidiA_details\n3. Commissions_BejaouiS_SaidiA_total\n4. rapport_commissions_CA\n5. etat_reglements_factures\n6. etat_factures_reglees100%_période_saisie\n7. Commissions_BejaouiS_periode_saisie\n8. Commissions_SaidiA_periode_saisie")
        root.destroy()




root = Tk()
root.title("Entrer date debut et date fin")
root.geometry("450x200")

label1 = Label(root, text="Veuillez entrer une date début ", font=("Arial", 13))
label2 = Label(root, text="Veuillez entrer une date fin ", font=("Arial", 13))

label1.place(x=10, y=10)
label2.place(x=10, y=90)

e1 = Entry(root)
e1.place(x=240, y = 10)

button1 = Button(root, text="Cliquer", command=deb_fun)
button1.place(x=375, y=10)

e2 = Entry(root, state='disabled')
e2.place(x=220, y=90)

button2 = Button(root, text="Confirmer", state='disabled', command=fin_fun)
button2.place(x=375, y=90)

root.mainloop()

#***OUTPUT 1***

#TABLEAU 1

data = pd.read_excel('MS_M_CalculComms.xlsx', header = 0)

del data['CO_No']

data2 = pd.read_excel('MSMARINE_CalculComms.xlsx', header = 0)

data = data.append(data2, ignore_index = True) #get all the data

data.drop_duplicates(subset = ["DOC"], keep="first", inplace=True) #drop duplicates
data.sort_values(by=['AN'])
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)
data.to_excel("data.xlsx") #get the duplicates free table on excel

#TABLEAU 2

ran = 2025 - 2016 + 1
ind = range(2016, 2016 + ran)
d = {	
    'Bejaoui Sahbi':[0.0] * ran,
    'Sahbi Bejaoui REV':[0.0] * ran,
    'Saidi Abdelkarim':[0.0] * ran,
    'Abdelkarim Saidi REV':[0.0] * ran  
}
df = pd.DataFrame(d, columns = ['Bejaoui Sahbi' , 'Sahbi Bejaoui REV', 'Saidi Abdelkarim' , 'Abdelkarim Saidi REV'], index = list(ind))

rpr = list(df.columns)
for y in ind:
	for rep in rpr:
		summ = 0
		for index, row in data.iterrows():
			if (row['AN'] == y and row['Representant'] == rep): 
				summ = summ + row['Total HT']
		df.at[y, rep] = summ

df.to_excel("CA_BejaouiS_SaidiA.xlsx") #CA


#TABLEAU 3
input2_BS = pd.read_excel('Objectifs_BejaouiS.xlsx', header = 0)

input2_BS = input2_BS.fillna(0)
input2_BS['CA'] = [0.0] * ran
input2_BS['Bejaoui Sahbi'] = [0.0] * ran
input2_BS['Sahbi Bejaoui REV'] = [0.0] * ran

BS = df["Bejaoui Sahbi"]
SB_REV = df["Sahbi Bejaoui REV"]

input2_SA = pd.read_excel('Objectifs_SaidiA.xlsx', header = 0)

input2_SA = input2_SA.fillna(0)
input2_SA['CA'] = [0.0] * ran
input2_SA['Saidi Abdelkarim'] = [0.0] * ran
input2_SA['Abdelkarim Saidi REV'] = [0.0] * ran

SA = df["Saidi Abdelkarim"]
AS_REV = df["Abdelkarim Saidi REV"]

x = 0
for index, row in df.iterrows():
	input2_BS.at[x, 'CA'] = row[rpr[0]] + row[rpr[1]]
	input2_BS.at[x, 'Bejaoui Sahbi'] = row[rpr[0]]
	input2_BS.at[x, 'Sahbi Bejaoui REV'] = row[rpr[1]]
	input2_SA.at[x, 'CA'] = row[rpr[2]] + row[rpr[3]]
	input2_SA.at[x, 'Saidi Abdelkarim'] = row[rpr[2]]
	input2_SA.at[x, 'Abdelkarim Saidi REV'] = row[rpr[3]]
	x = x + 1	
#print('CA SAHBI\n\n')
#print(input2_BS)		

table3 = df.copy()
table3.insert(2, "Bejaoui Sahbi EX", [0.0] * ran, True)
table3.insert(5, "Saidi Abdelkarim EX", [0.0] * ran, True)

x = 2016
for index, row in input2_BS.iterrows():
	if (row['Bejaoui Sahbi'] > row['objectif min']):
		table3.at[x, 'Bejaoui Sahbi'] = (row['Bejaoui Sahbi'] - row['objectif min']) * row['% VD']
		table3.at[x, 'Sahbi Bejaoui REV'] = row['Sahbi Bejaoui REV'] * row['% VR']
	else:
		table3.at[x, 'Bejaoui Sahbi'] = 0
		table3.at[x, 'Sahbi Bejaoui REV'] = 0
	if (row['CA'] > row['objectif excellence']):
		table3.at[x, 'Bejaoui Sahbi EX'] = row['CA'] * row['% excellence']	
	else:
		table3.at[x, 'Bejaoui Sahbi EX'] = 0
	x = x + 1

z = 2016
for index, row in input2_SA.iterrows():
	if (row['Saidi Abdelkarim'] > row['objectif min']):
		table3.at[z, 'Saidi Abdelkarim'] = (row['Saidi Abdelkarim'] - row['objectif min']) * row['% VD']
		table3.at[z, 'Abdelkarim Saidi REV'] = row['Abdelkarim Saidi REV'] * row['% VR']
	else:
		table3.at[z, 'Saidi Abdelkarim'] = 0
		table3.at[z, 'Abdelkarim Saidi REV'] = 0
	if (row['CA'] > row['objectif excellence']):
		table3.at[z, 'Saidi Abdelkarim EX'] = row['CA'] * row['% excellence']	
	else:
		table3.at[z, 'Saidi Abdelkarim EX'] = 0
	z = z + 1

table3.to_excel("Commissions_BejaouiS_SaidiA_details.xlsx") #tableau 3

# TABLEAU 4
f = {
    'Bejaoui Sahbi':[0.0] * ran,
    'Saidi Abdelkarim':[0.0] * ran  
}
table4 = pd.DataFrame(f, columns = ['Bejaoui Sahbi' , 'Saidi Abdelkarim'], index=list(ind))

for index, row in table3.iterrows():
	table4.at[index, 'Bejaoui Sahbi'] = row['Bejaoui Sahbi'] + row['Sahbi Bejaoui REV'] + row['Bejaoui Sahbi EX']
	table4.at[index, 'Saidi Abdelkarim'] = row['Saidi Abdelkarim'] + row['Abdelkarim Saidi REV'] + row['Saidi Abdelkarim EX']

table4.to_excel("Commissions_BejaouiS_SaidiA_total.xlsx")

#TABLEAU 5

table5 = table4.copy()
table4['CA BS'] = [0.0] * ran
table4['CA SA'] = [0.0] * ran

for index, row in df.iterrows():
	table4.at[index, 'CA BS'] = row[rpr[0]] + row[rpr[1]]
	table4.at[index, 'CA SA'] = row[rpr[2]] + row[rpr[3]]

for index, row in table4.iterrows():
	if(row['CA BS'] != 0):
		table5.at[index, 'Bejaoui Sahbi'] = row['Bejaoui Sahbi'] / row['CA BS']
	else:
		table5.at[index, 'Bejaoui Sahbi'] = 0
	if(row['CA SA'] != 0):
		table5.at[index, 'Saidi Abdelkarim'] = row['Saidi Abdelkarim'] / row['CA SA']
	else:
		table5.at[index, 'Saidi Abdelkarim'] = 0


table5.to_excel("rapport_commissions_CA.xlsx")


#***OUTPUT 2***

#TABLEAU A

data = pd.read_excel('MS_M_CalculComms.xlsx', header = 0)
del data['CO_No']

data2 = pd.read_excel('MSMARINE_CalculComms.xlsx', header = 0)

data = data.append(data2, ignore_index = True)
data.sort_values(by=['AN'])

pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

data = data.sort_values('Date_Regl')

data['Montant_Reglement'] = data.groupby(['DOC'])['Montant_Reglement'].transform('sum')

data = data.drop_duplicates('DOC', keep='last')

data['RAP'] = data['Total TTC'] - data['Montant_Reglement']
data['RAP'] = data['RAP'].apply(np.floor)

data.to_excel("etat_reglements_factures.xlsx")


#TABLEAU B


for index, row in data.iterrows():
	if (row['Date_Regl'] < date_deb or row['Date_Regl'] > date_fin):
		data.drop(index, inplace=True)

data.drop(data[data['RAP'] > 0].index, inplace = True)

data.to_excel("etat_factures_reglees100%_période_saisie.xlsx")

#TABLEAU C && D

orig_data = pd.read_excel('data.xlsx', header = 0)

table4 = pd.read_excel('Commissions_BejaouiS_SaidiA_total.xlsx', header = 0)

Obj_BS = pd.read_excel('Objectifs_BejaouiS.xlsx', header = 0, converters={'% VD':str,'% VR':str, '% excellence':str})

Obj_BS.set_index('AN', inplace=True, drop=True)

del Obj_BS['charges']

Obj_SA = pd.read_excel('Objectifs_SaidiA.xlsx', header = 0, converters={'% VD':str,'% VR':str, '% excellence':str})

Obj_SA.set_index('AN', inplace=True, drop=True)

del Obj_SA['charges']

ran = 2025 - 2016 + 1
ind = range(2016, 2016 + ran)
d = {	
    'factures HT année':[0.0] * ran,
    'factures TTC année':[0.0] * ran,
    'factures réglées période':[0.0] * ran,
    '% règlement':[0.0] * ran, 
    'commissions année':[0.0] * ran,
    'commission période':[0.0] * ran,
    'commission RAP':[0.0] * ran
}
BS = pd.DataFrame(d, columns = ['factures HT année', 'factures TTC année', 'factures réglées période', '% règlement', 'commissions année', 'commission période', 'commission RAP'], index = list(ind))

for y in ind:
    HT = 0
    TTC = 0
    for index, row in orig_data.iterrows():
        if ((row['AN'] == y) and (row['Representant'] == 'Bejaoui Sahbi' or row['Representant'] == 'Sahbi Bejaoui REV')):
            HT = HT + row['Total HT']
            TTC = TTC + row['Total TTC']
    regl = 0
    PR = 0
    for index, row in data.iterrows():
        if ((row['AN'] == y) and (row['Representant'] == 'Bejaoui Sahbi' or row['Representant'] == 'Sahbi Bejaoui REV')):
            regl = regl + row['Total TTC']
    BS.at[y, 'factures réglées période'] = regl
    if (TTC == 0):
        PR = 0
    else:
        PR = (regl / TTC) * 100
    BS.at[y, '% règlement'] = PR
    
    BS.at[y, 'factures HT année'] = HT
    BS.at[y, 'factures TTC année'] = TTC
     

SA = pd.DataFrame(d, columns = ['factures HT année', 'factures TTC année', 'factures réglées période', '% règlement', 'commissions année', 'commission période', 'commission RAP'], index = list(ind))

for y in ind:
    HT = 0
    TTC = 0
    for index, row in orig_data.iterrows():
        if ((row['AN'] == y) and (row['Representant'] == 'Saidi Abdelkarim' or row['Representant'] == 'Abdelkarim Saidi REV')):
            HT = HT + row['Total HT']
            TTC = TTC + row['Total TTC']
    SA.at[y, 'factures HT année'] = HT
    SA.at[y, 'factures TTC année'] = TTC
    regl = 0
    PR = 0
    for index, row in data.iterrows():
        if ((row['AN'] == y) and (row['Representant'] == 'Saidi Abdelkarim' or row['Representant'] == 'Abdelkarim Saidi REV')):
            regl = regl + row['Total TTC']
    SA.at[y, 'factures réglées période'] = regl
    if (TTC == 0):
        PR = 0
    else:
        PR = (regl / TTC) * 100
    SA.at[y, '% règlement'] = PR


#print(table4)

BS['com %'] = [0.0] * ran
SA['com %'] = [0.0] * ran
x = 2016
for index, row in table4.iterrows():
	BS.at[x, 'com %'] = row['Bejaoui Sahbi']
	SA.at[x, 'com %'] = row['Saidi Abdelkarim']
	x = x + 1

#print(BS)

CAn = 0
CP = 0
RAP = 0

y = 2016
for index, row in BS.iterrows():
    CAn = row['com %']
    BS.at[y, 'commissions année'] = CAn
    
    PR = row['% règlement']
    CP = (PR / 100) * CAn
    BS.at[y, 'commission période'] = CP
    
    RAP = CAn - CP
    BS.at[y, 'commission RAP'] = RAP

    y = y + 1

y = 2016
for index, row in SA.iterrows():
    CAn = row['com %']
    SA.at[y, 'commissions année'] = CAn
    
    PR = row['% règlement']
    CP = (PR / 100) * CAn
    SA.at[y, 'commission période'] = CP
    
    RAP = CAn - CP
    SA.at[y, 'commission RAP'] = RAP

    y = y + 1

BS = BS.append(BS.sum().rename('Total'))

SA = SA.append(SA.sum().rename('Total'))

del BS['com %']
del SA['com %']

Obj_BS = pd.concat([Obj_BS, BS], axis=1)

Obj_BS['% règlement'] = pd.Series(["{0:.1f}%".format(val) for val in Obj_BS['% règlement']], index = Obj_BS.index)

Obj_BS['% VD'] = pd.Series(["{0:.1f}%".format(float(val) * 100) for val in Obj_BS['% VD']], index = Obj_BS.index)

Obj_BS['% VR'] = pd.Series(["{0:.1f}%".format(float(val) * 100) for val in Obj_BS['% VR']], index = Obj_BS.index)

Obj_BS['% excellence'] = pd.Series(["{0:.1f}%".format(float(val) * 100) for val in Obj_BS['% excellence']], index = Obj_BS.index)

Obj_SA = pd.concat([Obj_SA, SA], axis=1)

Obj_SA['% règlement'] = pd.Series(["{0:.1f}%".format(val) for val in Obj_SA['% règlement']], index = Obj_SA.index)

Obj_SA['% VD'] = pd.Series(["{0:.1f}%".format(float(val) * 100) for val in Obj_SA['% VD']], index = Obj_SA.index)

Obj_SA['% VR'] = pd.Series(["{0:.1f}%".format(float(val) * 100) for val in Obj_SA['% VR']], index = Obj_SA.index)

Obj_SA['% excellence'] = pd.Series(["{0:.1f}%".format(float(val) * 100) for val in Obj_SA['% excellence']], index = Obj_SA.index)

Obj_BS.to_excel("Obj_BS.xlsx", float_format="%.0f")

wb = load_workbook("Obj_BS.xlsx")

ws1 = wb.active
ws2 = wb.create_sheet("new_BS")

start_row = 1
start_col = 1

for row in ws1.iter_rows(min_row=start_row):
    for cell in row:
        # print(cell.value)
        ws2.cell(row = start_row+3, column = start_col, value=cell.value) # start_row - 2 will assign the value to the same column up 2 rows
        start_col += 1 # increment the column, for use of destination sheet
    start_row += 1 # increment the row, for use of destination sheet
    start_col = 1 # reset to first column after row processing

#std=wb.get_sheet_by_name('Sheet1')

wb.active = ws2

ws2['A1'] = "Date debut:"
ws2['A2'] = date_deb.strftime('%d/%m/%y')

ws2['B1'] = "Date fin:"
ws2['B2'] = date_fin.strftime('%d/%m/%y')

ws2['J15'] = None

for row in ws2['D8:F15']:
    for cell in row:
        cell.value = None

del wb['Sheet1']

wb.save("Commissions_BejaouiS_periode_saisie.xlsx")

Obj_SA.to_excel("Obj_SA.xlsx", float_format="%.0f")

wb = load_workbook("Obj_SA.xlsx")

ws1 = wb.active
ws2 = wb.create_sheet("new_SA")

start_row = 1
start_col = 1

for row in ws1.iter_rows(min_row=start_row):
    for cell in row:
        # print(cell.value)
        ws2.cell(row = start_row+3, column = start_col, value=cell.value) # start_row - 2 will assign the value to the same column up 2 rows
        start_col += 1 # increment the column, for use of destination sheet
    start_row += 1 # increment the row, for use of destination sheet
    start_col = 1 # reset to first column after row processing

#std=wb.get_sheet_by_name('Sheet1')

wb.active = ws2

ws2['A1'] = "Date debut:"
ws2['A2'] = date_deb.strftime('%d/%m/%y')

ws2['B1'] = "Date fin:"
ws2['B2'] = date_fin.strftime('%d/%m/%y')

ws2['J15'] = None

for row in ws2['D8:F15']:
    for cell in row:
        cell.value = None

del wb['Sheet1']

wb.save("Commissions_SaidiA_periode_saisie.xlsx")


